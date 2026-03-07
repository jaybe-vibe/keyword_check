"""
Excel 리포트 생성기 - openpyxl 기반

3개 시트:
1. 키워드 목록 - 키워드별 검색량, 경쟁도, 분류
2. 상세 콘텐츠 - 키워드/블록별 개별 콘텐츠 (유형, 추천유형, 추천)
3. 연관 키워드 - 함께 많이 찾는 연관검색어
"""

import re
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from models import KeywordResult, BlockType, EXCLUDED_BLOCKS
from classifier import (
    analyze_keyword, get_content_type, is_recent,
)


# 스타일 상수
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
RECOMMEND_FILL = PatternFill(start_color="C6EFCE", fill_type="solid")
NOT_RECOMMEND_FILL = PatternFill(start_color="FFC7CE", fill_type="solid")


def _get_block_display_name(block) -> str:
    """블록의 표시 이름 결정. h2 헤더가 있으면 사용, 없으면 유형으로 대체."""
    name = block.block_name
    if name and not re.match(r'^[a-zA-Z0-9_]+$', name):
        return name
    return block.block_type.value


class ExcelReportGenerator:
    """Excel 분석 리포트 생성"""

    def __init__(self):
        self.wb = Workbook()

    def generate(self, results: dict, output_path: str) -> str:
        """
        전체 Excel 리포트 생성.

        Args:
            results: {keyword: KeywordResult} 딕셔너리
            output_path: 저장할 파일 경로

        Returns:
            저장된 파일 경로
        """
        self._create_keyword_sheet(results)
        self._create_detail_sheet(results)
        self._create_related_sheet(results)

        # 기본 빈 시트 제거
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # 디렉토리 생성
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        return output_path

    def _create_keyword_sheet(self, results: dict):
        """시트1: 키워드 목록 (검색량, 클릭수, 클릭률, 광고수, 분류 등)"""
        ws = self.wb.create_sheet("키워드 목록", 0)

        headers = [
            "키워드", "PC 검색량", "모바일 검색량", "총 검색량",
            "경쟁도",
            "월평균 PC클릭수", "월평균 모바일클릭수",
            "월평균 PC클릭률(%)", "월평균 모바일클릭률(%)",
            "월평균 노출 광고수",
            "분류", "스마트블록 수", "크롤링 시간", "비고",
        ]
        self._write_header_row(ws, headers)

        for row_idx, (keyword, result) in enumerate(results.items(), 2):
            if not isinstance(result, KeywordResult):
                continue
            ws.cell(row=row_idx, column=1, value=keyword)
            ws.cell(row=row_idx, column=2, value=result.search_volume_pc)
            ws.cell(row=row_idx, column=3, value=result.search_volume_mobile)
            ws.cell(row=row_idx, column=4, value=result.search_volume_total)
            ws.cell(row=row_idx, column=5, value=result.competition)
            ws.cell(row=row_idx, column=6, value=result.avg_pc_clicks)
            ws.cell(row=row_idx, column=7, value=result.avg_mobile_clicks)
            ws.cell(row=row_idx, column=8, value=result.avg_pc_ctr)
            ws.cell(row=row_idx, column=9, value=result.avg_mobile_ctr)
            ws.cell(row=row_idx, column=10, value=result.avg_ad_count)
            ws.cell(row=row_idx, column=11, value=result.classification)
            ws.cell(row=row_idx, column=12, value=len(result.smart_blocks))
            ws.cell(row=row_idx, column=13, value=(
                result.crawled_at.strftime("%Y-%m-%d %H:%M") if result.crawled_at else ""
            ))
            ws.cell(row=row_idx, column=14, value=result.error)

        self._auto_fit_columns(ws)

    def _create_detail_sheet(self, results: dict):
        """시트2: 키워드/블록별 상세 콘텐츠

        - 유형: URL 기반 분류 (블로그/인플루언서/카페/지식인)
        - 추천유형: 해당 키워드에서 조건 충족하는 유형 (카페/블로그 등)
        - 추천: 날짜가 최근 3개월 이내이면 추천
        """
        ws = self.wb.create_sheet("상세 콘텐츠")

        headers = [
            "키워드", "유형", "추천유형", "블록 이름", "순위",
            "제목", "URL", "출처", "날짜", "추천",
        ]
        self._write_header_row(ws, headers)

        row_idx = 2
        for keyword, result in results.items():
            if not isinstance(result, KeywordResult):
                continue

            # 키워드별 추천유형 계산
            analysis = analyze_keyword(result)
            recommended_label = "/".join(analysis["recommended_types"]) if analysis["recommended_types"] else ""

            global_rank = 0  # 키워드별 전체 순위

            for block in result.smart_blocks:
                if block.block_type in EXCLUDED_BLOCKS:
                    continue

                block_display = _get_block_display_name(block)

                for item in block.items:
                    global_rank += 1

                    # URL 기반 유형 분류
                    content_type = get_content_type(item.url)

                    # 날짜 기반 추천
                    recent = is_recent(item.date)
                    recommend = "추천" if recent is True else ""

                    ws.cell(row=row_idx, column=1, value=keyword)
                    ws.cell(row=row_idx, column=2, value=content_type)
                    ws.cell(row=row_idx, column=3, value=recommended_label)
                    ws.cell(row=row_idx, column=4, value=block_display)
                    ws.cell(row=row_idx, column=5, value=global_rank)
                    ws.cell(row=row_idx, column=6, value=item.title)
                    ws.cell(row=row_idx, column=7, value=item.url)
                    ws.cell(row=row_idx, column=8, value=item.source)
                    ws.cell(row=row_idx, column=9, value=item.date)

                    # 추천 셀 색상
                    rec_cell = ws.cell(row=row_idx, column=10, value=recommend)
                    if recent is True:
                        rec_cell.fill = RECOMMEND_FILL
                    elif recent is False:
                        rec_cell.fill = NOT_RECOMMEND_FILL
                    rec_cell.alignment = Alignment(horizontal="center")

                    row_idx += 1

        self._auto_fit_columns(ws)

    def _create_related_sheet(self, results: dict):
        """시트3: 연관 키워드 (함께 많이 찾는)"""
        ws = self.wb.create_sheet("연관 키워드")

        headers = ["원본 키워드", "연관 키워드"]
        self._write_header_row(ws, headers)

        row_idx = 2
        for keyword, result in results.items():
            if not isinstance(result, KeywordResult):
                continue
            for related in result.related_keywords:
                ws.cell(row=row_idx, column=1, value=keyword)
                ws.cell(row=row_idx, column=2, value=related)
                row_idx += 1

        self._auto_fit_columns(ws)

    def _write_header_row(self, ws, headers: list[str]):
        """헤더 행 작성 (스타일 포함)"""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

    def _auto_fit_columns(self, ws):
        """컬럼 너비 자동 조정"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    # 한글은 약 2배 너비
                    val_str = str(cell.value)
                    korean_count = sum(1 for c in val_str if ord(c) > 127)
                    length = len(val_str) + korean_count
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)
